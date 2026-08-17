from flask import Flask, render_template, request, redirect, url_for, jsonify, flash, session, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import os, json

app = Flask(__name__)
app.config['SECRET_KEY'] = 'expense-only-secret-2024'

# PostgreSQL on Railway, SQLite locally
db_url = os.environ.get('DATABASE_URL', '')
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = db_url or 'sqlite:///expenses.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
if db_url:
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'connect_args': {'connect_timeout': 10},
        'pool_pre_ping': True,
        'pool_recycle': 300,
    }

db = SQLAlchemy(app)

CATEGORIES_TREE = [
    {'name': 'الإسكان', 'name_en': 'Housing', 'emoji': '🏠', 'subs': [
        ('أسطوانات الغاز','Gas Cylinders','🔥'), ('الإيجار','Rent','🏢'), ('الخدمات','Utilities','💡'),
        ('الرهن العقاري','Mortgage','🏡'), ('الكهرباء','Electricity','⚡'), ('المياه','Water','💧'),
        ('صيانة المنزل','Home Maintenance','🔧'), ('ضرائب الممتلكات','Property Tax','📜'), ('مياه الصهريج','Water Tank','🚚'),
    ]},
    {'name': 'الأطفال والعائلة', 'name_en': 'Family & Kids', 'emoji': '👨‍👩‍👧', 'subs': [
        ('أنشطة مدرسية','School Activities','🏫'), ('تحويلات دعم عائلي','Family Support','🤝'),
        ('رعاية الأطفال','Childcare','🧸'), ('مستلزمات الأطفال','Baby Supplies','🍼'), ('مصروف','Allowance','💰'),
    ]},
    {'name': 'الأعمال والعمل', 'name_en': 'Business & Work', 'emoji': '🧾', 'subs': [
        ('الإعلانات والتسويق','Advertising','📢'), ('الخدمات القانونية','Legal Services','⚖️'),
        ('المحاسبة والضرائب','Accounting & Tax','🗃️'), ('النطاقات والاستضافة','Hosting & Domains','🌐'),
        ('تدريب ومؤتمرات','Training & Conferences','🎤'), ('تراخيص البرمجيات','Software Licenses','🧩'),
        ('خدمات مهنية','Professional Services','🤝'), ('لوازم المكتب','Office Supplies','📎'),
    ]},
    {'name': 'التأمين', 'name_en': 'Insurance', 'emoji': '🛡️', 'subs': [
        ('تأمين الأجهزة','Device Insurance','📱'), ('تأمين الحياة','Life Insurance','❤️'),
        ('تأمين السفر','Travel Insurance','✈️'), ('تأمين السيارة','Car Insurance','🚗'),
        ('تأمين المنزل','Home Insurance','🏠'),
    ]},
    {'name': 'الترفيه', 'name_en': 'Entertainment', 'emoji': '🎭', 'subs': [
        ('استراحات ومزارع','Retreats & Farms','🏕️'), ('الأفلام والبث','Movies & Streaming','🎥'),
        ('الألعاب','Games','🎮'), ('الحفلات والفعاليات','Events & Concerts','🎟️'),
        ('الرياضة واللياقة البدنية','Sports & Fitness','🤸'), ('الكتب والمجلات','Books & Magazines','📚'),
        ('المدن الترفيهية والمعالم','Theme Parks & Attractions','🎡'), ('حجز الملاعب','Court Booking','⚽'),
    ]},
    {'name': 'التسوق', 'name_en': 'Shopping', 'emoji': '🛍️', 'subs': [
        ('الأثاث','Furniture','🛋️'), ('الأحذية والإكسسوارات','Shoes & Accessories','👟'),
        ('الألعاب','Toys','🧸'), ('الإلكترونيات','Electronics','📱'),
        ('التسوق الإلكتروني','Online Shopping','🛒'), ('الجمال والعناية الشخصية','Beauty & Care','💄'),
        ('العبايات والدشاديش','Traditional Clothes','👘'), ('القرطاسية','Stationery','✏️'),
        ('المجوهرات والساعات','Jewelry & Watches','💎'), ('المعدات الرياضية','Sports Gear','🏀'),
        ('الملابس','Clothing','👗'), ('المنزل والمطبخ','Home & Kitchen','🔍'),
        ('الهدايا','Gifts','🎁'), ('رسوم الشحن والتوصيل','Shipping & Delivery','📦'),
    ]},
    {'name': 'التعليم', 'name_en': 'Education', 'emoji': '📚', 'subs': [
        ('الأنشطة اللامنهجية','Extracurricular','🥇'), ('الدروس الخصوصية','Private Tutoring','📖'),
        ('الدورات عبر الإنترنت','Online Courses','💻'), ('الشهادات','Certifications','⭐'),
        ('الكتب واللوازم','Books & Supplies','📚'), ('رسوم التعليم','Tuition Fees','🏫'),
    ]},
    {'name': 'الجهات الحكومية والرسوم', 'name_en': 'Government & Fees', 'emoji': '🏛️', 'subs': [
        ('التأشيرات والإقامة','Visas & Residency','📋'), ('الجمارك والرسوم','Customs & Duties','📦'),
        ('السجل التجاري والتراخيص','Business Registration','📋'), ('تجديد الهوية/الجواز','ID / Passport Renewal','🪪'),
        ('تصاريح العمل','Work Permits','🪪'), ('تصديق وتوثيق المستندات','Document Attestation','✏️'),
        ('رخصة القيادة','Driving License','🚦'), ('رسوم البلدية','Municipal Fees','📰'),
        ('رسوم قضائية وقانونية','Court Fees','⚖️'), ('مخالفات المرور','Traffic Fines','🚓'),
    ]},
    {'name': 'الحيوانات الأليفة', 'name_en': 'Pets', 'emoji': '🐾', 'subs': [
        ('الطب البيطري','Veterinary','🐶'), ('العناية بالحيوانات','Pet Grooming','✂️'),
        ('إيواء الحيوانات','Pet Boarding','🏠'), ('طعام الحيوانات','Pet Food','🦴'),
        ('مستلزمات الحيوانات','Pet Supplies','🧸'),
    ]},
    {'name': 'الخدمات الرقمية والاشتراكات', 'name_en': 'Digital & Subscriptions', 'emoji': '📦', 'subs': [
        ('أدوات الذكاء الاصطناعي','AI Tools','🤖'), ('أدوات المطور','Developer Tools','🧩'),
        ('اشتراكات أخرى','Other Subscriptions','🧾'), ('اشتراكات الألعاب','Gaming Subscriptions','🎮'),
        ('الإنترنت','Internet','🌐'), ('بث الموسيقى/الفيديو','Streaming','📺'),
        ('تخزين سحابي','Cloud Storage','☁️'), ('تطبيقات الإنتاجية','Productivity Apps','✅'),
        ('خطة الجوال','Mobile Plan','📱'), ('في بي إن والأمان','VPN & Security','🔒'),
        ('مشتريات داخل التطبيقات','In-App Purchases','📍'),
    ]},
    {'name': 'الزكاة والصدقات', 'name_en': 'Charity & Zakat', 'emoji': '🤲', 'subs': [
        ('زكاة','Zakat','🤲'), ('صدقة','Sadaqah','🤝'), ('الكفارات','Kaffarah','📿'), ('الأضحية','Udhiyah','🐑'),
    ]},
    {'name': 'السفر', 'name_en': 'Travel', 'emoji': '✈️', 'subs': [
        ('الجولات والأنشطة','Tours & Activities','🗺️'), ('الحج والعمرة','Hajj & Umrah','🕋'),
        ('الرحلات الجوية','Flights','✈️'), ('الفنادق','Hotels','🏨'),
        ('المطار والصالات','Airport & Lounges','🛫'), ('باقات السفر','Travel Packages','🏖️'),
        ('تأجير السيارات','Car Rental','🚗'), ('تأمين السفر','Travel Insurance','🛡️'),
        ('منتجعات وتذاكر يومية','Resorts & Day Passes','🏖️'),
    ]},
    {'name': 'الصحة والعافية', 'name_en': 'Health & Wellness', 'emoji': '🩺', 'subs': [
        ('الأدوية','Medications','💊'), ('الأسنان','Dental','🦷'), ('البصريات والنظارات','Opticals','👓'),
        ('التأمين الصحي','Health Insurance','🛡️'), ('التطعيمات','Vaccinations','💉'),
        ('الطب البديل (حجامة)','Alternative Medicine','🌿'), ('العلاج','Therapy','🛋️'),
        ('العلاج الطبيعي','Physiotherapy','🧴'), ('الفحوصات المخبرية','Lab Tests','🧪'),
        ('الفيتامينات والمكملات','Vitamins & Supplements','🌿'), ('المستشفى والعمليات','Hospital & Surgery','🏥'),
        ('زيارات الطبيب','Doctor Visits','🩺'),
    ]},
    {'name': 'الطعام وتناول الطعام', 'name_en': 'Food & Dining', 'emoji': '🍽️', 'subs': [
        ('البقالة','Groceries','🛒'), ('الطلبات الخارجية','Food Delivery','🍕'), ('العصائر','Juices','🧃'),
        ('المخابز والحلويات','Bakeries & Sweets','🎂'), ('المطاعم','Restaurants','🍽️'),
        ('المطاعم الفاخرة','Fine Dining','🍷'), ('المقاهي والقهوة','Cafes & Coffee','☕'),
        ('الوجبات الخفيفة','Snacks','🍪'), ('مياه معبأة','Bottled Water','💧'),
    ]},
    {'name': 'العقارات', 'name_en': 'Real Estate', 'emoji': '🏘️', 'subs': [
        ('شراء عقار','Property Purchase','🏠'), ('صيانة عقار','Property Maintenance','🔧'),
        ('رسوم وساطة','Brokerage Fees','📋'),
    ]},
    {'name': 'العمالة المنزلية', 'name_en': 'Domestic Staff', 'emoji': '🧺', 'subs': [
        ('راتب العمالة','Staff Salary','💵'), ('تذاكر وإقامة','Tickets & Iqama','✈️'),
        ('رسوم الاستقدام','Recruitment Fees','📋'),
    ]},
    {'name': 'العناية الشخصية', 'name_en': 'Personal Care', 'emoji': '🧴', 'subs': [
        ('الحلاقة والتجميل','Haircut & Beauty','✂️'), ('العطور ومستحضرات','Perfumes & Cosmetics','🧴'),
        ('النادي الرياضي','Gym','💪'), ('مستلزمات شخصية','Personal Supplies','🛁'),
    ]},
    {'name': 'المالية', 'name_en': 'Finance', 'emoji': '💰', 'subs': [
        ('أقساط قرض','Loan Installments','💳'), ('رسوم بنكية','Bank Fees','🏦'),
        ('استثمار','Investment','📈'), ('سداد ديون','Debt Repayment','💸'),
    ]},
    {'name': 'المركبات', 'name_en': 'Vehicles', 'emoji': '🚗', 'subs': [
        ('البترول والوقود','Fuel','⛽'), ('تأمين السيارة','Car Insurance','🛡️'),
        ('رسوم الطرق','Road Fees','🛣️'), ('فحص السيارة','Car Inspection','🔍'),
    ]},
    {'name': 'المناسبات والاحتفالات', 'name_en': 'Events & Celebrations', 'emoji': '🎊', 'subs': [
        ('أعراس وزواج','Weddings','💍'), ('مواليد وعقيقة','Newborn Celebration','👶'),
        ('تخرج','Graduation','🎓'), ('ديكور وترتيب','Decor & Setup','🎀'), ('ضيافة','Catering','🍽️'),
    ]},
    {'name': 'المواصلات', 'name_en': 'Transportation', 'emoji': '🚌', 'subs': [
        ('تاكسي وأوبر','Taxi & Uber','🚕'), ('باص ومواصلات عامة','Bus & Public Transit','🚌'),
        ('ركوب مشترك','Carpooling','🚗'), ('تذاكر','Tickets','🎫'),
    ]},
    {'name': 'النقد والصراف الآلي', 'name_en': 'Cash & ATM', 'emoji': '🏧', 'subs': [
        ('سحب نقدي','Cash Withdrawal','💵'), ('تحويل بنكي','Bank Transfer','🏦'),
        ('رسوم صراف','ATM Fees','💳'),
    ]},
    {'name': 'الهوايات', 'name_en': 'Hobbies', 'emoji': '🎨', 'subs': [
        ('أدوات هوايات','Hobby Tools','🎨'), ('دورات هوايات','Hobby Courses','📖'),
        ('مستلزمات رياضية','Sports Supplies','⚽'), ('نوادي واشتراكات','Clubs & Memberships','🏅'),
    ]},
    {'name': 'خدمات المنزل', 'name_en': 'Home Services', 'emoji': '🧰', 'subs': [
        ('الغسيل والتنظيف الجاف','Laundry & Dry Cleaning','🧺'), ('النقل والترحيل','Moving','📦'),
        ('أنظمة أمنية','Security Systems','🔔'), ('تنسيق حدائق','Landscaping','🌿'),
        ('تنظيف','Cleaning','🧹'), ('صيانة الأجهزة','Appliance Repair','🔧'),
        ('مكافحة الحشرات','Pest Control','✂️'),
    ]},
    {'name': 'عناية السيارة', 'name_en': 'Car Care', 'emoji': '🚗', 'subs': [
        ('إكسسوارات','Accessories','🧰'), ('الإطارات والعجلات','Tires & Wheels','⚙️'),
        ('البطارية وتغيير الزيت','Battery & Oil Change','🟢'), ('التسجيل والفحص','Registration & Inspection','🪪'),
        ('رسوم الطرق','Toll Fees','🛣️'), ('غسيل السيارة','Car Wash','🚿'), ('قطع الغيار','Spare Parts','⚙️'),
    ]},
    {'name': 'متفرقات', 'name_en': 'Miscellaneous', 'emoji': '🌀', 'subs': [
        ('التبرعات','Donations','🤝'), ('المشتريات المتنوعة','Miscellaneous Purchases','🛒'),
        ('النفقات الطائرة','Unexpected Expenses','❗'), ('أخرى','Other','📦'),
    ]},
]

# flat list for backward-compat lookups
CATEGORIES = [(s[0], s[0], s[-1]) for cat in CATEGORIES_TREE for s in cat['subs']]

def get_cat_emoji(name):
    for cat in CATEGORIES_TREE:
        for sub in cat['subs']:
            ar_name = sub[0]
            emoji   = sub[-1]   # last element is always the emoji
            if ar_name == name:
                return emoji
        if cat['name'] == name:
            return cat['emoji']
    return '📦'


class User(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(100), nullable=False, unique=True)
    pin_hash   = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    expenses   = db.relationship('Expense', backref='user', lazy=True, cascade='all, delete-orphan')
    budgets    = db.relationship('Budget', backref='user', lazy=True, cascade='all, delete-orphan')


class Expense(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    user_id     = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    amount      = db.Column(db.Float, nullable=False)
    category    = db.Column(db.String(50), nullable=False)
    description = db.Column(db.String(200))
    date        = db.Column(db.Date, default=date.today)
    month       = db.Column(db.Integer)
    year        = db.Column(db.Integer)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        if self.date:
            self.month = self.date.month
            self.year  = self.date.year


class Income(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    user_id     = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    amount      = db.Column(db.Float, nullable=False)
    source      = db.Column(db.String(200))
    description = db.Column(db.String(200))
    date        = db.Column(db.Date, default=date.today)
    month       = db.Column(db.Integer)
    year        = db.Column(db.Integer)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        if self.date:
            self.month = self.date.month
            self.year  = self.date.year


class MerchantMemory(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    user_id     = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    merchant    = db.Column(db.String(200), nullable=False)
    category    = db.Column(db.String(50),  nullable=False)
    updated_at  = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    __table_args__ = (db.UniqueConstraint('user_id', 'merchant'),)


class PushSub(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    endpoint   = db.Column(db.Text, nullable=False, unique=True)
    sub_json   = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Budget(db.Model):
    id      = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    month   = db.Column(db.Integer, nullable=False)
    year    = db.Column(db.Integer, nullable=False)
    amount  = db.Column(db.Float, nullable=False)


class NotifLog(db.Model):
    """تتبع الإشعارات المرسلة لتجنب التكرار"""
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    notif_type = db.Column(db.String(50), nullable=False)
    sent_date  = db.Column(db.Date, nullable=False, default=date.today)
    __table_args__ = (db.UniqueConstraint('user_id', 'notif_type', 'sent_date'),)


with app.app_context():
    try:
        db.create_all()
    except Exception as e:
        print(f"DB init warning: {e}")


def get_cat(key):
    emoji = get_cat_emoji(key)
    return (key, key, emoji)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# ── AUTH ──────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        name = request.form['name'].strip()
        pin  = request.form['pin'].strip()
        user = User.query.filter_by(name=name).first()
        if user and check_password_hash(user.pin_hash, pin):
            session['user_id']   = user.id
            session['user_name'] = user.name
            return redirect(url_for('index'))
        flash('الاسم أو الرمز غير صحيح ❌', 'error')
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form['name'].strip()
        pin  = request.form['pin'].strip()
        if len(pin) != 4 or not pin.isdigit():
            flash('الرمز يجب أن يكون 4 أرقام ❌', 'error')
            return render_template('register.html')
        if User.query.filter_by(name=name).first():
            flash('هذا الاسم مستخدم بالفعل ❌', 'error')
            return render_template('register.html')
        user = User(name=name, pin_hash=generate_password_hash(pin))
        db.session.add(user)
        db.session.commit()
        session['user_id']   = user.id
        session['user_name'] = user.name
        flash(f'مرحباً {name}! تم إنشاء حسابك 🎉', 'success')
        return redirect(url_for('index'))
    return render_template('register.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── MAIN ──────────────────────────────────────────
@app.route('/')
@login_required
def index():
    user_id = session['user_id']
    today   = date.today()
    month   = request.args.get('month', today.month, type=int)
    year    = request.args.get('year',  today.year,  type=int)
    search  = request.args.get('q', '').strip()

    query = Expense.query.filter_by(user_id=user_id, month=month, year=year)
    if search:
        query = query.filter(
            db.or_(
                Expense.description.ilike(f'%{search}%'),
                Expense.category.ilike(f'%{search}%')
            )
        )
    expenses = query.all()
    total    = sum(e.amount for e in expenses)

    # all expenses this month (without search filter) for stats
    all_expenses = Expense.query.filter_by(user_id=user_id, month=month, year=year).all()
    all_total    = sum(e.amount for e in all_expenses)

    by_cat = {}
    for e in all_expenses:
        by_cat[e.category] = by_cat.get(e.category, 0) + e.amount

    # budget
    budget = Budget.query.filter_by(user_id=user_id, month=month, year=year).first()
    budget_amount = budget.amount if budget else None
    budget_pct    = round(all_total / budget_amount * 100, 1) if budget_amount else None

    # prev month comparison
    prev_month = month - 1 if month > 1 else 12
    prev_year  = year if month > 1 else year - 1
    prev_total = db.session.query(db.func.sum(Expense.amount))\
                   .filter_by(user_id=user_id, month=prev_month, year=prev_year).scalar() or 0
    diff_pct = round((all_total - prev_total) / prev_total * 100, 1) if prev_total > 0 else None

    arabic_months = {
        1:'يناير',2:'فبراير',3:'مارس',4:'أبريل',
        5:'مايو',6:'يونيو',7:'يوليو',8:'أغسطس',
        9:'سبتمبر',10:'أكتوبر',11:'نوفمبر',12:'ديسمبر'
    }
    years_list = list(range(today.year - 2, today.year + 2))

    return render_template('index.html',
        expenses=expenses,
        total=total,
        all_total=all_total,
        by_cat=by_cat,
        categories=CATEGORIES,
        categories_tree=CATEGORIES_TREE,
        current_month=month,
        current_year=year,
        years_list=years_list,
        arabic_months=arabic_months,
        get_cat=get_cat,
        user_name=session.get('user_name'),
        search=search,
        budget_amount=budget_amount,
        budget_pct=budget_pct,
        diff_pct=diff_pct,
        prev_month=prev_month,
        prev_total=prev_total,
    )


@app.route('/add_expense', methods=['POST'])
@login_required
def add_expense():
    amount      = float(request.form['amount'])
    category    = request.form['category']
    description = request.form.get('description', '')
    entry_date  = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
    e = Expense(user_id=session['user_id'], amount=amount,
                category=category, description=description, date=entry_date)
    db.session.add(e)
    db.session.commit()
    flash('تمت إضافة المصروف ✅', 'success')
    return redirect(url_for('index', month=entry_date.month, year=entry_date.year))


@app.route('/edit_expense/<int:id>', methods=['POST'])
@login_required
def edit_expense(id):
    e = Expense.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    e.amount      = float(request.form['amount'])
    new_category  = request.form['category']
    e.description = request.form.get('description', '')
    entry_date    = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
    e.date  = entry_date
    e.month = entry_date.month
    e.year  = entry_date.year

    # حفظ في ذاكرة المتاجر إذا تغيّرت الفئة وفيه اسم متجر
    if new_category != e.category and e.description:
        merchant_key = e.description.strip().upper()
        mem = MerchantMemory.query.filter_by(
            user_id=session['user_id'], merchant=merchant_key).first()
        if mem:
            mem.category   = new_category
            mem.updated_at = datetime.utcnow()
        else:
            db.session.add(MerchantMemory(
                user_id  = session['user_id'],
                merchant = merchant_key,
                category = new_category
            ))

    e.category = new_category
    db.session.commit()
    flash('تم التعديل ✅', 'success')
    return redirect(url_for('index', month=e.month, year=e.year))


@app.route('/delete_expense/<int:id>')
@login_required
def delete_expense(id):
    e = Expense.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    month, year = e.month, e.year
    db.session.delete(e)
    db.session.commit()
    flash('تم الحذف ✅', 'info')
    return redirect(url_for('index', month=month, year=year))


@app.route('/set_budget', methods=['POST'])
@login_required
def set_budget():
    month  = int(request.form['month'])
    year   = int(request.form['year'])
    amount = float(request.form['budget_amount'])
    budget = Budget.query.filter_by(user_id=session['user_id'], month=month, year=year).first()
    if budget:
        budget.amount = amount
    else:
        budget = Budget(user_id=session['user_id'], month=month, year=year, amount=amount)
        db.session.add(budget)
    db.session.commit()
    flash('تم حفظ الميزانية ✅', 'success')
    return redirect(url_for('index', month=month, year=year))


@app.route('/income')
@login_required
def income_page():
    user_id = session['user_id']
    today   = date.today()
    month   = request.args.get('month', today.month, type=int)
    year    = request.args.get('year',  today.year,  type=int)

    incomes      = Income.query.filter_by(user_id=user_id, month=month, year=year).order_by(Income.date.desc()).all()
    total_income = sum(i.amount for i in incomes)
    all_expenses = Expense.query.filter_by(user_id=user_id, month=month, year=year).all()
    total_expense = sum(e.amount for e in all_expenses)
    balance      = total_income - total_expense

    prev_month = month - 1 if month > 1 else 12
    prev_year  = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year  = year if month < 12 else year + 1

    month_names = ['يناير','فبراير','مارس','أبريل','مايو','يونيو',
                   'يوليو','أغسطس','سبتمبر','أكتوبر','نوفمبر','ديسمبر']

    return render_template('income.html',
        incomes=incomes, total_income=total_income,
        total_expense=total_expense, balance=balance,
        month=month, year=year,
        month_name=month_names[month-1],
        prev_month=prev_month, prev_year=prev_year,
        next_month=next_month, next_year=next_year,
    )


@app.route('/add_income', methods=['POST'])
@login_required
def add_income():
    user_id = session['user_id']
    try:
        amount  = float(request.form['amount'])
        source  = request.form.get('source', '').strip() or 'دخل'
        date_str = request.form.get('date', str(date.today()))
        inc_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        flash('بيانات غير صحيحة', 'error')
        return redirect(url_for('income_page'))

    db.session.add(Income(user_id=user_id, amount=amount, source=source,
                          description=source, date=inc_date))
    db.session.commit()
    flash('تم إضافة الدخل', 'success')
    return redirect(url_for('income_page'))


@app.route('/delete_income/<int:income_id>')
@login_required
def delete_income(income_id):
    inc = Income.query.filter_by(id=income_id, user_id=session['user_id']).first_or_404()
    db.session.delete(inc)
    db.session.commit()
    return redirect(url_for('income_page'))


@app.route('/export_excel')
@login_required
def export_excel():
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('مكتبة التصدير غير متاحة', 'error')
        return redirect(url_for('index'))

    user_id = session['user_id']
    today   = date.today()
    month   = request.args.get('month', today.month, type=int)
    year    = request.args.get('year',  today.year,  type=int)

    expenses = Expense.query.filter_by(user_id=user_id, month=month, year=year)\
                            .order_by(Expense.date).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    arabic_months = {1:'يناير',2:'فبراير',3:'مارس',4:'أبريل',5:'مايو',6:'يونيو',
                     7:'يوليو',8:'أغسطس',9:'سبتمبر',10:'أكتوبر',11:'نوفمبر',12:'ديسمبر'}
    ws.title = f"{arabic_months[month]} {year}"

    # Header
    headers = ['التاريخ', 'الفئة', 'الوصف', 'المبلغ (ر.ع)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='C0392B')
        cell.alignment = Alignment(horizontal='center')

    for row, e in enumerate(expenses, 2):
        cat = get_cat(e.category)
        ws.cell(row=row, column=1, value=str(e.date))
        ws.cell(row=row, column=2, value=f"{cat[2]} {cat[1]}")
        ws.cell(row=row, column=3, value=e.description or '')
        ws.cell(row=row, column=4, value=round(e.amount, 3))

    # Total row
    total_row = len(expenses) + 2
    ws.cell(row=total_row, column=3, value='الإجمالي').font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=4, value=round(sum(e.amount for e in expenses), 3))
    total_cell.font = Font(bold=True, color='C0392B')

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"مصروفات_{arabic_months[month]}_{year}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/chart')
@login_required
def chart():
    user_id = session['user_id']
    today   = date.today()
    mode    = request.args.get('mode', 'monthly')
    year    = request.args.get('year',  today.year,  type=int)
    month   = request.args.get('month', today.month, type=int)

    if mode == 'daily':
        import calendar
        days_in_month = calendar.monthrange(year, month)[1]
        labels = [str(d) for d in range(1, days_in_month + 1)]
        data = []
        for d in range(1, days_in_month + 1):
            try:
                day_date = date(year, month, d)
            except ValueError:
                data.append(0)
                continue
            total = db.session.query(db.func.sum(Expense.amount))\
                      .filter(Expense.user_id == user_id, Expense.date == day_date).scalar() or 0
            data.append(round(total, 3))
    else:
        labels = ['يناير','فبراير','مارس','أبريل','مايو','يونيو',
                  'يوليو','أغسطس','سبتمبر','أكتوبر','نوفمبر','ديسمبر']
        data = []
        for m in range(1, 13):
            total = db.session.query(db.func.sum(Expense.amount))\
                      .filter_by(user_id=user_id, month=m, year=year).scalar() or 0
            data.append(round(total, 3))

    return jsonify({'labels': labels, 'data': data, 'mode': mode})


@app.route('/scan_receipt', methods=['POST'])
@login_required
def scan_receipt():
    import requests as http_requests
    from PIL import Image
    import io, json, re, base64

    api_key = os.environ.get('MISTRAL_API_KEY') or os.environ.get('GEMINI_API_KEY')
    if not api_key:
        return jsonify({'error': 'مفتاح AI غير موجود'}), 500

    if 'image' not in request.files:
        return jsonify({'error': 'لم يتم إرسال صورة'}), 400

    file = request.files['image']
    img_bytes = file.read()

    try:
        try:
            import pillow_heif
            pillow_heif.register_heif_opener()
        except ImportError:
            pass

        img = Image.open(io.BytesIO(img_bytes))
        if img.mode != 'RGB':
            img = img.convert('RGB')

        max_size = 1600
        if max(img.size) > max_size:
            ratio = max_size / max(img.size)
            img = img.resize((int(img.size[0]*ratio), int(img.size[1]*ratio)), Image.LANCZOS)

        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=85)
        jpeg_b64 = base64.b64encode(buf.getvalue()).decode('utf-8')

        prompt = (
            "You are a receipt reader. Look at this receipt image carefully.\n"
            "Reply with ONLY a raw JSON object, no markdown, no explanation:\n"
            '{"amount": <total amount as number>, '
            '"description": "<store name or description in Arabic>", '
            '"category": "<one of: food, groceries, coffee, petrol, carwash, carmaint, health, pharmacy, education, entertainment, clothing, utilities, internet, subscriptions, savings, gifts, travel, housing, other>", '
            '"date": "<date as YYYY-MM-DD or empty string>"}\n\n'
            "Category rules: coffee shop/cafe→coffee, gas station/fuel→petrol, restaurant→food, "
            "supermarket/grocery→groceries, pharmacy→pharmacy, hospital/clinic→health, "
            "clothes/shoes→clothing, electricity/water bill→utilities, carwash→carwash, "
            "car repair/parts→carmaint, otherwise→other. "
            "If amount unclear use 0. Write description in Arabic."
        )

        url = "https://api.mistral.ai/v1/chat/completions"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        mistral_payload = {
            "model": "pixtral-12b-2409",
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": f"data:image/jpeg;base64,{jpeg_b64}"}
                ]
            }],
            "max_tokens": 300
        }
        resp = http_requests.post(url, json=mistral_payload, headers=headers, timeout=30)
        resp.raise_for_status()
        result = resp.json()
        text = result['choices'][0]['message']['content'].strip()

        text = re.sub(r'```(?:json)?', '', text).strip()
        match = re.search(r'\{.*\}', text, re.DOTALL)
        if match:
            data = json.loads(match.group())
        else:
            data = json.loads(text)

        try:
            data['amount'] = float(str(data.get('amount', 0)).replace(',', '.'))
        except Exception:
            data['amount'] = 0

        return jsonify({'success': True, 'data': data})

    except Exception as e:
        return jsonify({'error': str(e), 'detail': 'scan_failed'}), 500




import re as _re
from flask import send_from_directory
import requests as _requests

def ai_categorize(merchant_name, user_id=None):
    """تصنيف ذكي: ذاكرة المتاجر → Gemini AI → كلمات مفتاحية"""

    # ١. تحقق من ذاكرة المتاجر أولاً
    if user_id:
        merchant_key = merchant_name.strip().upper()
        mem = MerchantMemory.query.filter_by(user_id=user_id, merchant=merchant_key).first()
        if mem:
            return mem.category

    api_key = os.environ.get('GEMINI_API_KEY') or os.environ.get('MISTRAL_API_KEY')
    if not api_key:
        return _fallback_categorize(merchant_name)

    categories = (
        "housing=السكن والإيجار, food=المطاعم والأكل, groceries=البقالة والسوبرماركت, "
        "coffee=القهوة والشاي والعصائر, petrol=البترول ومحطات الوقود, "
        "carwash=غسيل السيارة, carmaint=صيانة السيارة وقطع الغيار, "
        "health=الصحة والمستشفيات والعيادات, pharmacy=الصيدلية والأدوية, "
        "education=التعليم والكتب والدورات, entertainment=الترفيه والسينما والألعاب, "
        "clothing=الملابس والأحذية والأزياء, internet=الهاتف والإنترنت والاتصالات, "
        "subscriptions=الاشتراكات الرقمية نتفليكس سبوتيفاي, "
        "gifts=الهدايا, travel=السفر والفنادق والطيران, "
        "personal=العناية الشخصية الصالون الحلاقة, other=أخرى"
    )

    valid = [c[0] for c in CATEGORIES]

    prompt = (
        f"Merchant name from bank SMS: '{merchant_name}'\n"
        f"Choose ONE category key from this list:\n"
        f"housing, food, groceries, coffee, petrol, carwash, carmaint, "
        f"health, pharmacy, education, entertainment, clothing, internet, "
        f"subscriptions, gifts, travel, personal, other\n\n"
        f"Rules:\n"
        f"- coffee/cafe/tea/juice/drink → coffee\n"
        f"- restaurant/burger/pizza/shawarma/grill/مطعم → food\n"
        f"- lulu/carrefour/supermarket/hypermarket/market/grocery → groceries\n"
        f"- petrol/fuel/gas station/محطة → petrol\n"
        f"- pharmacy/drug/صيدلية → pharmacy\n"
        f"- hospital/clinic/medical/doctor → health\n"
        f"- salon/barber/spa/beauty → personal\n"
        f"- hotel/airline/travel/airport → travel\n"
        f"Reply with ONLY the key word. Nothing else."
    )

    def _extract_cat(text):
        text = text.strip().lower()
        # البحث عن أول كلمة مفتاحية صحيحة في الرد
        for word in _re.findall(r'[a-z]+', text):
            if word in valid:
                return word
        return None

    try:
        if os.environ.get('GEMINI_API_KEY'):
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={os.environ['GEMINI_API_KEY']}"
            resp = _requests.post(url, json={
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"maxOutputTokens": 15, "temperature": 0}
            }, timeout=6)
            if resp.ok:
                text = resp.json()['candidates'][0]['content']['parts'][0]['text']
                cat  = _extract_cat(text)
                if cat:
                    return cat

        if os.environ.get('MISTRAL_API_KEY'):
            resp = _requests.post("https://api.mistral.ai/v1/chat/completions", json={
                "model": "mistral-small-latest",
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 15, "temperature": 0
            }, headers={"Authorization": f"Bearer {os.environ['MISTRAL_API_KEY']}"}, timeout=6)
            if resp.ok:
                text = resp.json()['choices'][0]['message']['content']
                cat  = _extract_cat(text)
                if cat:
                    return cat

    except Exception:
        pass

    return _fallback_categorize(merchant_name)


def _fallback_categorize(name):
    """تصنيف احتياطي بالكلمات المفتاحية"""
    lower = name.lower()
    if any(w in lower for w in ['coffee', 'cafe', 'tea', 'juice', 'قهوة', 'شاي']):
        return 'coffee'
    if any(w in lower for w in ['restaurant', 'مطعم', 'burger', 'pizza', 'grill', 'shawarma']):
        return 'food'
    if any(w in lower for w in ['petrol', 'fuel', 'station', 'بترول', 'وقود']):
        return 'petrol'
    if any(w in lower for w in ['pharmacy', 'صيدلية', 'drug']):
        return 'pharmacy'
    if any(w in lower for w in ['lulu', 'carrefour', 'hypermarket', 'supermarket', 'بقالة']):
        return 'groceries'
    if any(w in lower for w in ['hospital', 'clinic', 'medical', 'مستشفى', 'عيادة']):
        return 'health'
    return 'other'

@app.route('/sw.js')
def sw():
    resp = send_from_directory('static', 'sw.js')
    resp.headers['Service-Worker-Allowed'] = '/'
    resp.headers['Cache-Control'] = 'no-cache'
    return resp

@app.route('/sms_webhook', methods=['POST'])
def sms_webhook():
    """
    iOS Shortcuts يرسل رسالة SMS من bankmuscat هنا تلقائياً.
    JSON body: { "message": "...", "username": "...", "pin": "..." }
    """
    data = request.get_json(silent=True) or {}
    message_body = data.get('message', '')
    username     = data.get('username', '')
    pin          = data.get('pin', '')

    if not message_body:
        return jsonify({'error': 'no message'}), 400

    is_debit  = 'تم خصم' in message_body
    is_credit = any(k in message_body for k in ['تم إيداع', 'تم اضافة', 'تم إضافة', 'تم استلام', 'تم قيد'])

    if not is_debit and not is_credit:
        return jsonify({'ignored': 'not a recognized bank message'}), 200

    # التحقق من المستخدم
    user = User.query.filter_by(name=username).first()
    if not user or not check_password_hash(user.pin_hash, str(pin)):
        return jsonify({'error': 'unauthorized'}), 401

    # استخراج المبلغ
    amount = None
    amt_match = _re.search(r'([\d,]+\.\d+)\s*(?:OMR|RO)|(?:OMR|RO)\s*([\d,]+\.\d+)', message_body, _re.IGNORECASE)
    if amt_match:
        try:
            raw = amt_match.group(1) or amt_match.group(2)
            amount = float(raw.replace(',', ''))
        except Exception:
            amount = None

    if not amount or amount <= 0:
        return jsonify({'error': 'could not parse amount', 'message': message_body}), 422

    if is_debit:
        desc_match  = _re.search(r'في\s+(?:[\d\w]+-)?(.+?)\s+بتاريخ', message_body)
        description = desc_match.group(1).strip() if desc_match else 'بنك مسقط - دفعة تلقائية'
        cat = ai_categorize(description, user_id=user.id)
        db.session.add(Expense(
            user_id=user.id, amount=amount, category=cat,
            description=description, date=date.today()
        ))
        db.session.commit()
        return jsonify({'success': True, 'type': 'expense', 'amount': amount, 'category': cat, 'description': description}), 200
    else:
        src_match = _re.search(r'من\s+(.+?)\s+(?:بتاريخ|في)', message_body)
        source = src_match.group(1).strip() if src_match else 'إيداع بنكي'
        db.session.add(Income(
            user_id=user.id, amount=amount, source=source,
            description=message_body[:200], date=date.today()
        ))
        db.session.commit()
        return jsonify({'success': True, 'type': 'income', 'amount': amount, 'source': source}), 200


@app.route('/api/ai_insights', methods=['GET'])
@login_required
def ai_insights():
    import calendar as cal_mod
    from collections import defaultdict

    user_id = session['user_id']
    now = date.today()

    # جلب آخر 3 أشهر من المصروفات
    expenses = Expense.query.filter(
        Expense.user_id == user_id,
        Expense.date >= date(now.year - (1 if now.month <= 3 else 0),
                            (now.month - 3) % 12 or 12,
                            1)
    ).all()

    if not expenses:
        return jsonify({'insights': [], 'tip': 'أضف مصاريف أولاً لتحصل على تحليلات ذكية!'})

    # إحصاءات
    by_weekday = defaultdict(float)   # 0=Mon..6=Sun
    by_category = defaultdict(float)
    by_week = defaultdict(float)
    daily_totals = defaultdict(float)
    current_month_exp = [e for e in expenses if e.month == now.month and e.year == now.year]
    prev_month = (now.month - 2) % 12 + 1
    prev_year  = now.year if now.month > 1 else now.year - 1
    prev_month_exp = [e for e in expenses if e.month == prev_month and e.year == prev_year]

    for e in expenses:
        by_weekday[e.date.weekday()] += e.amount
        by_category[e.category] += e.amount
        daily_totals[e.date] += e.amount

    # أيام الأسبوع بالعربية
    days_ar = ['الاثنين','الثلاثاء','الأربعاء','الخميس','الجمعة','السبت','الأحد']

    insights = []
    cat_map = {k: v for k, v, _ in CATEGORIES}

    # ١. أكثر يوم إنفاقاً
    if by_weekday:
        busiest_day = max(by_weekday, key=by_weekday.get)
        avg = by_weekday[busiest_day] / max(1, sum(1 for e in expenses if e.date.weekday() == busiest_day))
        insights.append({
            'icon': '📅',
            'title': f'يوم {days_ar[busiest_day]} أعلى أيامك إنفاقاً',
            'detail': f'متوسط {avg:,.0f} لكل {days_ar[busiest_day]}'
        })

    # ٢. أعلى فئة
    if by_category:
        top_cat = max(by_category, key=by_category.get)
        top_pct = by_category[top_cat] / sum(by_category.values()) * 100
        insights.append({
            'icon': '🏆',
            'title': f'{cat_map.get(top_cat, top_cat)} تستهلك {top_pct:.0f}٪ من إنفاقك',
            'detail': f'المجموع {by_category[top_cat]:,.0f}'
        })

    # ٣. مقارنة الشهر الحالي بالسابق
    curr_total = sum(e.amount for e in current_month_exp)
    prev_total = sum(e.amount for e in prev_month_exp)
    if prev_total > 0:
        diff_pct = (curr_total - prev_total) / prev_total * 100
        if abs(diff_pct) >= 5:
            direction = 'زاد' if diff_pct > 0 else 'انخفض'
            emoji = '📈' if diff_pct > 0 else '📉'
            insights.append({
                'icon': emoji,
                'title': f'إنفاقك هذا الشهر {direction} {abs(diff_pct):.0f}٪',
                'detail': f'الشهر الماضي: {prev_total:,.0f} | هذا الشهر: {curr_total:,.0f}'
            })

    # ٤. أعلى يوم إنفاق
    if daily_totals:
        peak_day = max(daily_totals, key=daily_totals.get)
        insights.append({
            'icon': '🔥',
            'title': f'أعلى يوم إنفاق كان {peak_day.strftime("%-d/%-m")}',
            'detail': f'{daily_totals[peak_day]:,.0f} في يوم واحد'
        })

    # ٥. متوسط يومي هذا الشهر
    if current_month_exp:
        days_passed = now.day
        daily_avg = curr_total / days_passed
        projected = daily_avg * cal_mod.monthrange(now.year, now.month)[1]
        insights.append({
            'icon': '🔮',
            'title': f'توقع نهاية الشهر: {projected:,.0f}',
            'detail': f'بناءً على متوسطك اليومي {daily_avg:,.0f}'
        })

    # توليد نصيحة AI
    api_key = os.environ.get('GEMINI_API_KEY')
    tip = None
    if api_key and by_category:
        try:
            import requests as _req, json as _json
            top_cats = sorted(by_category.items(), key=lambda x: -x[1])[:3]
            cats_text = '، '.join(f'{cat_map.get(c,c)} ({v:,.0f})' for c, v in top_cats)
            prompt = (
                f"المستخدم أنفق هذا الشهر على: {cats_text}. "
                f"المجموع {curr_total:,.0f}. "
                f"اكتب نصيحة ذكية واحدة مختصرة بالعربية (جملة واحدة فقط، بدون تحية)."
            )
            r = _req.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={api_key}",
                json={'contents': [{'parts': [{'text': prompt}]}]},
                timeout=8
            )
            if r.ok:
                tip = r.json()['candidates'][0]['content']['parts'][0]['text'].strip()
        except Exception:
            pass

    return jsonify({'insights': insights, 'tip': tip})


@app.route('/api/verify_pin', methods=['POST'])
@login_required
def verify_pin():
    data = request.get_json() or {}
    pin  = str(data.get('pin', ''))
    user = User.query.get(session['user_id'])
    if user and check_password_hash(user.pin_hash, pin):
        return jsonify({'ok': True})
    return jsonify({'error': 'رمز خاطئ'}), 401


@app.route('/api/change_pin', methods=['POST'])
@login_required
def change_pin():
    data    = request.get_json() or {}
    old_pin = str(data.get('old_pin', ''))
    new_pin = str(data.get('new_pin', ''))
    user    = User.query.get(session['user_id'])
    if not check_password_hash(user.pin_hash, old_pin):
        return jsonify({'error': 'الرمز الحالي غير صحيح'}), 400
    if len(new_pin) != 4 or not new_pin.isdigit():
        return jsonify({'error': 'الرمز الجديد يجب أن يكون 4 أرقام'}), 400
    user.pin_hash = generate_password_hash(new_pin)
    db.session.commit()
    return jsonify({'ok': True})


VAPID_PRIVATE = os.environ.get('VAPID_PRIVATE_KEY', '')
VAPID_PUBLIC  = os.environ.get('VAPID_PUBLIC_KEY', '')
VAPID_EMAIL   = 'mailto:admin@masarify.app'


@app.route('/api/vapid_public')
def vapid_public():
    return jsonify({'key': VAPID_PUBLIC})


@app.route('/api/subscribe_push', methods=['POST'])
@login_required
def subscribe_push():
    sub = request.get_json()
    if not sub or 'endpoint' not in sub:
        return jsonify({'error': 'invalid'}), 400
    existing = PushSub.query.filter_by(endpoint=sub['endpoint']).first()
    if existing:
        existing.sub_json = json.dumps(sub)
        existing.user_id  = session['user_id']
    else:
        db.session.add(PushSub(
            user_id  = session['user_id'],
            endpoint = sub['endpoint'],
            sub_json = json.dumps(sub)
        ))
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/unsubscribe_push', methods=['POST'])
@login_required
def unsubscribe_push():
    data = request.get_json() or {}
    PushSub.query.filter_by(user_id=session['user_id'], endpoint=data.get('endpoint','')).delete()
    db.session.commit()
    return jsonify({'ok': True})


def send_push(user_id, title, body):
    """إرسال إشعار لجميع أجهزة المستخدم"""
    if not VAPID_PRIVATE or not VAPID_PUBLIC:
        return
    try:
        from pywebpush import webpush, WebPushException
    except ImportError:
        return
    subs = PushSub.query.filter_by(user_id=user_id).all()
    for s in subs:
        try:
            webpush(
                subscription_info   = json.loads(s.sub_json),
                data                = json.dumps({'title': title, 'body': body}),
                vapid_private_key   = VAPID_PRIVATE,
                vapid_claims        = {'sub': VAPID_EMAIL}
            )
        except Exception:
            db.session.delete(s)
    db.session.commit()


@app.route('/api/check_budget_notify')
@login_required
def check_budget_notify():
    return smart_notify()


def _already_sent(user_id, notif_type, today):
    return NotifLog.query.filter_by(
        user_id=user_id, notif_type=notif_type, sent_date=today).first() is not None


def _mark_sent(user_id, notif_type, today):
    try:
        db.session.add(NotifLog(user_id=user_id, notif_type=notif_type, sent_date=today))
        db.session.commit()
    except Exception:
        db.session.rollback()


@app.route('/api/smart_notify')
@login_required
def smart_notify():
    """نظام إشعارات ذكي شامل — يُستدعى عند فتح التطبيق"""
    import calendar as cal_mod
    from collections import defaultdict

    user_id = session['user_id']
    today   = date.today()
    results = []

    # ── ١. تنبيه الميزانية (80 / 90 / 100٪) ──
    budget = Budget.query.filter_by(user_id=user_id, month=today.month, year=today.year).first()
    if budget:
        total = db.session.query(db.func.sum(Expense.amount))\
                  .filter_by(user_id=user_id, month=today.month, year=today.year).scalar() or 0
        pct = total / budget.amount * 100

        if pct >= 100 and not _already_sent(user_id, 'budget_100', today):
            send_push(user_id, '🚨 تجاوزت ميزانيتك!',
                      f'صرفت {total:.3f} من {budget.amount:.3f} ر.ع هذا الشهر')
            _mark_sent(user_id, 'budget_100', today)
            results.append('budget_100')
        elif pct >= 90 and not _already_sent(user_id, 'budget_90', today):
            send_push(user_id, '⚠️ 90٪ من ميزانيتك!',
                      f'بقي لك {budget.amount - total:.3f} ر.ع فقط هذا الشهر')
            _mark_sent(user_id, 'budget_90', today)
            results.append('budget_90')
        elif pct >= 80 and not _already_sent(user_id, 'budget_80', today):
            send_push(user_id, '🔔 80٪ من ميزانيتك',
                      f'صرفت {pct:.0f}٪ — بقي {budget.amount - total:.3f} ر.ع')
            _mark_sent(user_id, 'budget_80', today)
            results.append('budget_80')

    # ── ٢. تقرير أسبوعي (كل أحد) ──
    if today.weekday() == 6 and not _already_sent(user_id, 'weekly', today):
        week_start = today - __import__('datetime').timedelta(days=6)
        week_total = db.session.query(db.func.sum(Expense.amount))\
                       .filter(Expense.user_id == user_id,
                               Expense.date >= week_start,
                               Expense.date <= today).scalar() or 0
        from collections import defaultdict
        by_cat = defaultdict(float)
        for e in Expense.query.filter(Expense.user_id == user_id,
                                      Expense.date >= week_start,
                                      Expense.date <= today).all():
            by_cat[e.category] += e.amount
        cat_map = {k: v for k, v, _ in CATEGORIES}
        top = max(by_cat, key=by_cat.get) if by_cat else None
        top_name = cat_map.get(top, top) if top else ''
        send_push(user_id, f'📊 تقرير الأسبوع',
                  f'صرفت {week_total:.3f} ر.ع هذا الأسبوع'
                  + (f' · أعلى فئة: {top_name}' if top_name else ''))
        _mark_sent(user_id, 'weekly', today)
        results.append('weekly')

    # ── ٣. نصيحة يومية بـ Gemini ──
    api_key = os.environ.get('GEMINI_API_KEY')
    if api_key and not _already_sent(user_id, 'daily_tip', today):
        try:
            import requests as _req
            expenses_month = Expense.query.filter_by(
                user_id=user_id, month=today.month, year=today.year).all()
            if len(expenses_month) >= 5:
                from collections import defaultdict
                by_cat = defaultdict(float)
                for e in expenses_month:
                    by_cat[e.category] += e.amount
                cat_map = {k: v for k, v, _ in CATEGORIES}
                top_cats = sorted(by_cat.items(), key=lambda x: -x[1])[:2]
                cats_text = ' و'.join(cat_map.get(c, c) for c, _ in top_cats)
                total_m = sum(e.amount for e in expenses_month)
                days_left = cal_mod.monthrange(today.year, today.month)[1] - today.day
                prompt = (
                    f"المستخدم أنفق {total_m:.0f} ر.ع هذا الشهر، أعلى فئاته: {cats_text}، "
                    f"بقي {days_left} يوم في الشهر. "
                    f"اكتب نصيحة مالية ذكية قصيرة جداً بالعربية (جملة واحدة فقط)."
                )
                r = _req.post(
                    f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={api_key}",
                    json={'contents': [{'parts': [{'text': prompt}]}]},
                    timeout=8
                )
                if r.ok:
                    tip = r.json()['candidates'][0]['content']['parts'][0]['text'].strip()
                    send_push(user_id, '💡 نصيحة اليوم', tip)
                    _mark_sent(user_id, 'daily_tip', today)
                    results.append('daily_tip')
        except Exception:
            pass

    return jsonify({'ok': True, 'sent': results})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5051))
    app.run(host='0.0.0.0', debug=True, port=port)
